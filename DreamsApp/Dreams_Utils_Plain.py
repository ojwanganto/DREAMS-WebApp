from django.core.files.storage import default_storage
import openpyxl as xl
import xlrd
import traceback
from Dreams_Excel_Mapping import *
from string import Template
import datetime
from django.db import connection
from django.conf import settings
from openpyxl.utils.exceptions import *
import os


class DreamsEnrollmentExcelTemplateRenderer(object):

    document_path = ''
    workbook = ''

    def __init__(self):
        pass

    def create_tmp_file(self, excel):
        try:
            with open(default_storage.path('tmp/enrollment.xlsx'), 'wb+') as destination:
                for chunk in excel.chunks():
                    destination.write(chunk)
            destination.closed
            self.document_path = default_storage.path('tmp/enrollment.xlsx')
            return default_storage.path('tmp/enrollment.xlsx')
        except:
            print 'Could not create a temp file'
            return

    def get_document_path(self):
        return self.document_path

    def validate_excel_document(self):
        # check to see if Main Database sheet exists
        # TODO define and implement more checks

        wb = xl.load_workbook(self.get_document_path(), read_only=True, keep_vba=False)
        self.workbook = wb
        if 'Main Database' in wb.get_sheet_names():
            return True
        else:
            return False

    def get_sheet_names(self):
        if self.validate_excel_document():
            return xl.load_workbook(self.get_document_path(), read_only=True, keep_vba=False).get_sheet_names()
        else:
            return 'Invalid Sheets'

    def excel_enrollment_data(self):
        if self.validate_excel_document():
            try:
                sql_statement = self.generate_insert_SQL(6, 50)
                # print sql_statement
                self.execute_SQL_Query(sql_statement)

            except Exception as e:
                tb = traceback.format_exc()
                print 'There was an error ' + tb
            return

    def generate_insert_SQL(self, start, end):
        if self.validate_excel_document():
            try:
                print 'The Excel Database is Valid'
                print "Now processing the file.........."

                sql_statement = """
INSERT INTO DreamsApp_client
(first_name, middle_name, last_name, dreams_id, dss_id_number, implementing_partner_id, date_of_birth, date_of_enrollment, marital_status_id, phone_number,
county_of_residence_id, sub_county_id, ward_id, informal_settlement, village, landmark, guardian_name, relationship_with_guardian, guardian_phone_number, guardian_national_id, is_date_of_birth_estimated, verification_document_id, verification_doc_no )
VALUES """

                rangeList = range(start, end)
                lastNumber = rangeList[-1]
                row_qry = ''
                for rowid in range(start, end):

                    if rowid != lastNumber:
                        row_qry += self.generate_row_insert_SQL(rowid) + ', \n'
                    else:
                        row_qry += " " + self.generate_row_insert_SQL(rowid)

                return Template('$insert $values').substitute(insert=sql_statement, values=row_qry)

            except Exception as e:
                tb = traceback.format_exc()
                print 'There was an error ' + tb
            return

    def get_enrollment_sheet_openpyxl(self):
        return xl.load_workbook(self.get_document_path(), read_only=True, data_only=True).get_sheet_by_name('Main Database')

    def get_enrollment_sheet(self):
        return self.get_wb().sheet_by_name('Main Database')

    def get_wb(self):
        return xlrd.open_workbook(self.get_document_path(), on_demand=True)

    def get_row_data(self, row_num):

        sheet = self.get_enrollment_sheet_openpyxl()
        cols = self.openpyxl_demographic_columns()

        # map fields that should mirror webapp's

        ver_doc = str(sheet.cell(row=row_num, column=cols.get('verification_doc')).value)
        ver_doc = ExcelDreamsMapping.verification_document().get(ver_doc)

        marital_status = str(sheet.cell(row=row_num, column=cols.get('marital_status')).value)
        marital_status = ExcelDreamsMapping.marital_status_codes().get(marital_status)

        county = str(sheet.cell(row=row_num, column=cols.get('county')).value)
        county = ExcelDreamsMapping.county().get(county)

        sub_county = str(sheet.cell(row=row_num, column=cols.get('sub_county')).value)
        sub_county = ExcelDreamsMapping.sub_county().get(sub_county)

        ward = str(sheet.cell(row=row_num, column=cols.get('ward')).value)
        ward = ExcelDreamsMapping.ward_by_name().get(ward)

        dob_raw = sheet.cell(row=row_num, column=cols.get('dob')).value
        if dob_raw is not None:
            dob_raw = datetime.datetime.strftime(dob_raw, '%Y-%m-%d')

        doe_raw = sheet.cell(row=row_num, column=cols.get('date_of_enrollment')).value
        if doe_raw is not None:
            doe_raw = datetime.datetime.strftime(doe_raw, '%Y-%m-%d')

        row_values = {
            'serial_No': str(sheet.cell(row=row_num, column=cols.get('serial_No')).value),
            'IP': str(sheet.cell(row=row_num, column=cols.get('IP')).value),
            'first_name': str(sheet.cell(row=row_num, column=cols.get('first_name')).value),
            'middle_name': str(sheet.cell(row=row_num, column=cols.get('middle_name')).value),
            'last_name': str(sheet.cell(row=row_num, column=cols.get('last_name')).value),
            'dob': dob_raw,
            'verification_doc': ver_doc,
            'verification_doc_other': str(sheet.cell(row=row_num, column=cols.get('verification_doc_other')).value),
            'verification_doc_no': str(sheet.cell(row=row_num, column=cols.get('verification_doc_no')).value),
            'date_of_enrollment': doe_raw,
            'marital_status': marital_status,
            'client_phone_no': str(sheet.cell(row=row_num, column=cols.get('client_phone_no')).value),
            'county': county,
            'sub_county': sub_county,
            'ward': ward,
            'informal_settlement': str(sheet.cell(row=row_num, column=cols.get('informal_settlement')).value),
            'village': str(sheet.cell(row=row_num, column=cols.get('village')).value),
            'land_mark': str(sheet.cell(row=row_num, column=cols.get('land_mark')).value),
            'dreams_id': str(sheet.cell(row=row_num, column=cols.get('dreams_id')).value),
            'dss_id': str(sheet.cell(row=row_num, column=cols.get('dss_id')).value),
            'caregiver_first_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_first_name')).value),
            'caregiver_middle_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_middle_name')).value),
            'caregiver_last_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_last_name')).value),
            'caregiver_relationship': str(sheet.cell(row=row_num, column=cols.get('caregiver_relationship')).value),
            'caregiver_relationship_other': str(sheet.cell(row=row_num, column=cols.get('caregiver_relationship_other')).value),
            'caregiver_phone_no': str(sheet.cell(row=row_num, column=cols.get('caregiver_phone_no')).value),
            'caregiver_id_no': str(sheet.cell(row=row_num, column=cols.get('caregiver_id_no')).value),
        }
        return row_values


    def generate_row_insert_SQL(self, row_num):

        row_data = self.get_row_data(row_num)
        caregiver_fname = row_data.get('caregiver_first_name')
        caregiver_mname = row_data.get('caregiver_middle_name')
        caregiver_lname = row_data.get('caregiver_last_name')

        coded_rel = row_data.get('caregiver_relationship')
        other_rel = row_data.get('caregiver_relationship_other')
        date_of_birth_estimated = 0

        # django app has no field for other relationship. check which one has value

        if coded_rel != 'None' or other_rel != 'None':
            if other_rel != 'None':
                caregiver_relationship = other_rel
            else:
                caregiver_relationship = coded_rel
        else:
            caregiver_relationship = ''

        ver_doc = row_data.get('verification_doc')
        ver_doc_other = row_data.get('verification_doc_other')

        # TODO: Handle other verification documents. It is not currently in the Client's model. Code therefore picks only coded documents

        guardian_name = (caregiver_fname + " " if caregiver_fname != 'None' else "") + (caregiver_mname + " " if caregiver_mname != 'None' else "") + (caregiver_lname if caregiver_lname != 'None' else "")

        values_template = Template(
            '("$first_name", "$middle_name", "$last_name", "$dreams_id", "$dss_id_number", "$implementing_partner_id", "$date_of_birth", "$date_of_enrollment", "$marital_status_id", "$phone_number", \
    "$county_of_residence_id", "$sub_county_id", "$ward_id", "$informal_settlement", "$village", "$landmark", "$guardian_name", "$relationship_with_guardian", "$guardian_phone_number", "$guardian_national_id", "$is_date_of_birth_estimated", "$verification_doc_id", "$verification_doc_no" )').safe_substitute(
            first_name=row_data.get('first_name') if row_data.get('first_name') !='None' else '',
            middle_name=row_data.get('middle_name') if row_data.get('middle_name') !='None' else '',
            last_name=row_data.get('last_name') if row_data.get('last_name') !='None' else '',
            dreams_id=row_data.get('dreams_id'),
            dss_id_number=row_data.get('dss_id') if row_data.get('dss_id') !='None' else '',
            implementing_partner_id=row_data.get('IP') if row_data.get('IP') !='None' else '',
            date_of_birth=row_data.get('dob') if row_data.get('dob') !='None' else '0000-00-00',
            date_of_enrollment=row_data.get('date_of_enrollment') if row_data.get('date_of_enrollment') !='None' else '0000-00-00',
            marital_status_id=row_data.get('marital_status') if row_data.get('marital_status') !='None' else '',
            phone_number=row_data.get('client_phone_no') if row_data.get('client_phone_no') !='None' else '',
            county_of_residence_id=row_data.get('county'),
            sub_county_id=row_data.get('sub_county'),
            ward_id=row_data.get('ward'),
            informal_settlement=row_data.get('informal_settlement') if row_data.get('informal_settlement') !='None' else '',
            village=row_data.get('village') if row_data.get('village') !='None' else '',
            landmark=row_data.get('land_mark') if row_data.get('land_mark') !='None' else '',
            guardian_name=guardian_name,
            relationship_with_guardian=caregiver_relationship,
            guardian_phone_number=row_data.get('caregiver_phone_no') if row_data.get('caregiver_phone_no') !='None' else '',
            guardian_national_id=row_data.get('caregiver_id_no') if row_data.get('caregiver_id_no') !='None' else '',
            is_date_of_birth_estimated=date_of_birth_estimated,
            verification_doc_id=row_data.get('verification_doc'),
            verification_doc_no=row_data.get('verification_doc_no')
        )

        return values_template

    def openpyxl_demographic_columns(self):
        demographics = {
            'serial_No': 1,
            'IP': 3,
            'first_name': 4,
            'middle_name': 5,
            'last_name': 6,
            'dob': 7,
            'verification_doc': 8,
            'verification_doc_other': 9,
            'verification_doc_no': 10,
            'date_of_enrollment': 11,
            'marital_status': 16,
            'client_phone_no': 17,
            'county': 18,
            'sub_county': 19,
            'ward': 20,
            'informal_settlement': 22,
            'village': 23,
            'land_mark': 24,
            'dreams_id': 25,
            'dss_id': 26,
            'caregiver_first_name': 27,
            'caregiver_middle_name': 28,
            'caregiver_last_name': 29,
            'caregiver_relationship': 30,
            'caregiver_relationship_other': 31,
            'caregiver_phone_no': 32,
            'caregiver_id_no': 33,
        }

        return demographics

    def dump_SQL(self):
        partner_id = 8
        return """ SELECT
            client_id,
            /*first_name,
            middle_name,
            last_name,*/
            date_of_birth,
            verification_document_id,
            /*verification_doc_no,*/
            date_of_enrollment,
            /*phone_number,*/
            dss_id_number,
            informal_settlement,
            village,
            landmark,
            /*dreams_id,
            guardian_name,*/
            relationship_with_guardian,
            /*guardian_phone_number,
            guardian_national_id,*/
            county_of_residence_id,
            implementing_partner_id,
            marital_status_id,
            sub_county_id,
            ward_id,
            ward_name,
            sub_county_code,
            sub_county_name,
            county_code,
            county_name,
            head_of_household_id,
            head_of_household_other,
            age_of_household_head,
            is_father_alive,
            is_mother_alive,
            is_parent_chronically_ill,
            main_floor_material_id,
            main_floor_material_other,
            main_roof_material_id,
            main_roof_material_other,
            main_wall_material_id,
            main_wall_material_other,
            source_of_drinking_water_id,
            source_of_drinking_water_other,
            no_of_adults,
            no_of_females,
            no_of_males,
            no_of_children,
            currently_in_ct_program_id,
            current_ct_program,
            ever_enrolled_in_ct_program_id,
            ever_missed_full_day_food_in_4wks_id,
            has_disability_id,
            no_of_days_missed_food_in_4wks_id,
            no_of_people_in_household,
            age_at_first_sexual_encounter,
            sex_partners_in_last_12months,
            age_of_last_partner_id,
            age_of_second_last_partner_id,
            age_of_third_last_partner_id,
            ever_had_sex_id,
            has_sexual_partner_id,
            know_last_partner_hiv_status_id,
            know_second_last_partner_hiv_status_id,
            know_third_last_partner_hiv_status_id,
            last_partner_circumcised_id,
            received_money_gift_for_sex_id,
            second_last_partner_circumcised_id,
            third_last_partner_circumcised_id,
            used_condom_with_last_partner_id,
            used_condom_with_second_last_partner_id,
            used_condom_with_third_last_partner_id,
            no_of_biological_children,
            anc_facility_name,
            known_fp_method_other,
            current_fp_method_other,
            reason_not_using_fp_other,
            current_anc_enrollment_id,
            current_fp_method_id,
            currently_pregnant_id,
            currently_use_modern_fp_id,
            fp_methods_awareness_id,
            has_biological_children_id,
            reason_not_using_fp_id,
            drug_abuse_last_12months_other,
            drug_used_last_12months_other,
            drug_abuse_last_12months_id,
            frequency_of_alcohol_last_12months_id,
            produced_alcohol_last_12months_id,
            used_alcohol_last_12months_id,
            dreams_program_other,
            gbv_help_provider_other,
            preferred_gbv_help_provider_other,
            economic_threat_ever_id,
            economic_threat_last_3months_id,
            humiliated_ever_id,
            humiliated_last_3months_id,
            insulted_ever_id,
            insulted_last_3months_id,
            knowledge_of_gbv_help_centres_id,
            physical_violence_ever_id,
            physical_violence_last_3months_id,
            physically_forced_other_sex_acts_ever_id,
            physically_forced_other_sex_acts_last_3months_id,
            physically_forced_sex_ever_id,
            physically_forced_sex_last_3months_id,
            seek_help_after_gbv_id,
            threatened_for_sexual_acts_ever_id,
            threatened_for_sexual_acts_last_3months_id,
            threats_to_hurt_ever_id,
            threats_to_hurt_last_3months_id,
            current_school_name,
            current_class,
            current_school_level_other,
            current_education_supporter_other,
            reason_not_in_school_other,
            dropout_class,
            life_wish_other,
            current_income_source_other,
            banking_place_other,
            banking_place_id,
            current_income_source_id,
            current_school_level_id,
            current_school_type_id,
            currently_in_school_id,
            dropout_school_level_id,
            has_savings_id,
            last_time_in_school_id,
            life_wish_id,
            reason_not_in_school_id,
            care_facility_enrolled,
            reason_not_in_hiv_care_other,
            reason_never_tested_for_hiv_other,
            enrolled_in_hiv_care_id,
            ever_tested_for_hiv_id,
            knowledge_of_hiv_test_centres_id,
            last_test_result_id,
            period_last_tested_id,
            reason_not_in_hiv_care_id
  from flat_dreams_enrollment WHERE implementing_partner_id = %d  """ %(partner_id)

    def execute_SQL_Query(self, sql):
        cursor = connection.cursor()
        try:
            cursor.execute(sql)

            print cursor.rowcount
        except Exception as e:
            print 'There was an Error running the query\n'
            traceback.format_exc()

        return

    def get_export_rows(self, ip_list_str):
        sql = self.dump_SQL()
        cursor = connection.cursor()
        #ip_list = (1, 2)
        try:

            ip_tuple_l = ip_list_str.split(",")
            if len(ip_tuple_l) > 1:
                eval_listt = eval(ip_list_str)
                ip_list = tuple(eval_listt)
                cursor.execute("SELECT * FROM flat_dreams_enrollment WHERE implementing_partner_id IN %s ", [ip_list])
            else:
                ip_list = int(ip_list_str)
                cursor.execute("SELECT * FROM flat_dreams_enrollment WHERE implementing_partner_id = %s ", [ip_list])

            print "Query was successful"
            columns = [col[0] for col in cursor.description]
            return [
                dict(zip(columns, row))
                for row in cursor.fetchall()
                ]
        except Exception as e:
            print 'There was an Error running the query\n'
            traceback.format_exc()

        return

    def load_workbook(self):
        DREAMS_TEMPLATE_PLAIN = os.path.join(settings.BASE_DIR, 'templates/excel_template/dreams_export.xlsx')
        try:
            wb = xl.load_workbook(DREAMS_TEMPLATE_PLAIN)
            return wb
        except InvalidFileException as e:
            traceback.format_exc()

    def load_workbook_template(self):
        DREAMS_TEMPLATE_PLAIN = os.path.join(settings.BASE_DIR, 'templates/excel_template/sample_template.xlsx')
        try:
            wb = xl.load_workbook(DREAMS_TEMPLATE_PLAIN)
            return wb
        except InvalidFileException as e:
            traceback.format_exc()

    def prepare_excel_doc(self, ip_list_str):

        try:

            wb = self.load_workbook()
            refined_sheet = wb.get_sheet_by_name('enrollment_refined')
            print "Starting DB Query! ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db_data = self.get_export_rows(ip_list_str)
            print "Finished DB Query. Rendering Now. ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            i = 1
            for row in db_data:
                i += 1
                self.map_demographics(refined_sheet, i, row)
                self.map_individual_and_household(refined_sheet, i, row)
                self.map_sexuality(refined_sheet, i, row)
                self.map_reproductive_health(refined_sheet, i, row)
                self.map_drug_use(refined_sheet, i, row)
                self.map_education_and_employment(refined_sheet, i, row)
                self.map_gbv(refined_sheet, i, row)
                self.map_program_participation(refined_sheet, i, row)
                self.map_hiv_testing(refined_sheet, i, row)

            wb.save('dreams_enrollment_interventions.xlsx')
            print "Completed rendering excel ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return wb
        except InvalidFileException as e:
            traceback.format_exc()
        except ReadOnlyWorkbookException as e:
            traceback.format_exc()

        except SheetTitleException as e:
            traceback.format_exc()
        return

    def map_demographics(self, ws, i, row):
        cols = {
            'implementing_partner_id': 2,
            'IP_Code': 3,
            #'first_name': 4,
            #'middle_name': 5,
            #'last_name': 6,
            'date_of_birth': 7,
            'verification_document_id': 8,
            'verification_doc_other': 9,
            #'verification_doc_no': 10,
            'date_of_enrollment': 11,
            'marital_status_id': 16,
            #'phone_number': 17,
            'county_name': 18,
            'sub_county_name': 19,
            'ward_id': 21,
            'ward_name': 20,
            'informal_settlement': 22,
            'village': 23,
            'land_mark': 24,
            'dreams_id': 25,
            'dss_id_number': 26,
            # 'caregiver_first_name': 25,
            # 'caregiver_middle_name': 26,
            # 'caregiver_last_name': 27,
            'relationship_with_guardian': 30,
            'caregiver_relationship_other': 31
            #'guardian_phone_number': 32,
            #'guardian_national_id': 33,
        }

        for k, v in cols.items():
            if k == 'implementing_partner_id':
                val = row.get(k)
                if val is not None:
                    partner = self.map_implementing_partner().get(val)
                    ws.cell(row=i, column=v, value=partner)
            elif k == 'IP_Code':
                val = row.get('implementing_partner_id')
                ws.cell(row=i, column=v, value=val)
            elif k == 'verification_document_id':
                val = row.get(k)
                if val is not None:
                    item = self.map_verification_document().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'marital_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.map_marital_status_codes().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_individual_and_household(self, ws, i, row):
        cols = {
            'head_of_household_id': 34,
            'head_of_household_other': 35,
            'age_of_household_head': 36,
            'is_father_alive': 37,
            'is_mother_alive': 38,
            'is_parent_chronically_ill': 39,
            'main_floor_material_id': 40,
            'main_floor_material_other': 41,
            'main_roof_material_id': 42,
            'main_roof_material_other': 43,
            'main_wall_material_id': 44,
            'main_wall_material_other': 45,
            'source_of_drinking_water_id': 46,
            'source_of_drinking_water_other': 47,
            'no_of_people_in_household': 57,
            'no_of_adults': 60,
            'no_of_females': 58,
            'no_of_males': 59,
            'no_of_children': 61,
            'currently_in_ct_program_id': 63,
            'current_ct_program': 64,
            'ever_enrolled_in_ct_program_id': 62,
            'ever_missed_full_day_food_in_4wks_id': 48,
            'has_disability_id': 50,
            'no_of_days_missed_food_in_4wks_id': 49,
            'disability_types': 51
        }

        for k, v in cols.items():
            if k == 'head_of_household_id':
                val = row.get(k)
                if val is not None:
                    item = self.headOfHouseHoldDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_disability_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'disability_types':
                val = row.get(k)
                if val is not None:
                    dts = val.split(",")
                    for d in dts:
                        if d == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_disability_type().get(d), value='Yes')
            elif k == 'is_father_alive':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'is_mother_alive':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'is_parent_chronically_ill':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'main_floor_material_id':
                val = row.get(k)
                if val is not None:
                    item = self.floorMaterialDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'main_roof_material_id':
                val = row.get(k)
                if val is not None:
                    item = self.roofMaterialDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'main_wall_material_id':
                val = row.get(k)
                if val is not None:
                    item = self.wallMaterialDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'source_of_drinking_water_id':
                val = row.get(k)
                if val is not None:
                    item = self.drinkingWaterSourceDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_in_ct_program_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_enrolled_in_ct_program_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_missed_full_day_food_in_4wks_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'no_of_days_missed_food_in_4wks_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_sexuality(self, ws, i, row):
        cols = {
            'age_at_first_sexual_encounter': 108,
            'sex_partners_in_last_12months': 109,
            'age_of_last_partner_id': 111,
            'age_of_second_last_partner_id': 112,
            'age_of_third_last_partner_id': 113,
            'ever_had_sex_id': 107,
            'has_sexual_partner_id': 110,
            'know_last_partner_hiv_status_id': 117,
            'know_second_last_partner_hiv_status_id': 118,
            'know_third_last_partner_hiv_status_id': 119,
            'last_partner_circumcised_id': 114,
            'received_money_gift_for_sex_id': 123,
            'second_last_partner_circumcised_id': 115,
            'third_last_partner_circumcised_id': 116,
            'used_condom_with_last_partner_id': 120,
            'used_condom_with_second_last_partner_id': 121,
            'used_condom_with_third_last_partner_id': 122
        }
        for k, v in cols.items():
            if k == 'age_of_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.relativeAgeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'age_of_second_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.relativeAgeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'age_of_third_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.relativeAgeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_had_sex_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_sexual_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'know_last_partner_hiv_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'know_second_last_partner_hiv_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'know_third_last_partner_hiv_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'last_partner_circumcised_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'second_last_partner_circumcised_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'third_last_partner_circumcised_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'received_money_gift_for_sex_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_condom_with_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_condom_with_second_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_condom_with_third_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_reproductive_health(self, ws, i, row):
        cols = {
            'no_of_biological_children': 125,
            'anc_facility_name': 128,
            'known_fp_method_other': 134,
            'current_fp_method_other': 139,
            'reason_not_using_fp_other': 141,
            'current_anc_enrollment_id': 127,
            'current_fp_method_id': 138,
            'currently_pregnant_id': 126,
            'currently_use_modern_fp_id': 137,
            'fp_methods_awareness_id': 129,
            'has_biological_children_id': 124,
            'reason_not_using_fp_id': 140,
            'known_fp_methods': 130
        }
        for k, v in cols.items():
            if k == 'known_fp_methods':
                val = row.get(k)
                if val is not None:
                    fpm = val.split(",")
                    for m in fpm:
                        if m == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_fp_method().get(m), value='Yes')
            elif k == 'current_anc_enrollment_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_fp_method_id':
                val = row.get(k)
                if val is not None:
                    item = self.familyPlanningDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_pregnant_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_use_modern_fp_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'fp_methods_awareness_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_biological_children_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'reason_not_using_fp_id':
                val = row.get(k)
                if val is not None:
                    item = self.reasonNoInFPDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_drug_use(self, ws, i, row):
        cols = {
            # 'dr.drug_abuse_last_12months_other': 180,
            'drug_used_last_12months_other': 191,
            'drug_abuse_last_12months_id': 184,
            'frequency_of_alcohol_last_12months_id': 183,
            'produced_alcohol_last_12months_id': 194,
            'used_alcohol_last_12months_id': 182,
            'drugs_used_in_last_12_months': 185
        }
        for k, v in cols.items():
            if k == 'drugs_used_in_last_12_months':
                val = row.get(k)
                if val is not None:
                    drugs = val.split(",")
                    for d in drugs:
                        if d == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_drugs().get(d), value='Yes')
            elif k == 'frequency_of_alcohol_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'drug_abuse_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'produced_alcohol_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_alcohol_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_program_participation(self, ws, i, row):
        cols = {
            'dreams_program_other': 203,
            'programmes_enrolled': 195
        }
        for k, v in cols.items():
            if k == 'programmes_enrolled':
                val = row.get(k)
                if val is not None:
                    pgs = val.split(",")
                    for p in pgs:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_dreams_programmes().get(p), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_gbv(self, ws, i, row):
        cols = {
            'gbv_help_provider_other': 169,
            'preferred_gbv_help_provider_other': 181,
            'economic_threat_ever_id': 148,
            'economic_threat_last_3months_id': 149,
            'humiliated_ever_id': 142,
            'humiliated_last_3months_id': 143,
            'insulted_ever_id': 146,
            'insulted_last_3months_id': 147,
            'knowledge_of_gbv_help_centres_id': 170,
            'physical_violence_ever_id': 150,
            'physical_violence_last_3months_id': 151,
            'physically_forced_other_sex_acts_ever_id': 154,
            'physically_forced_other_sex_acts_last_3months_id': 155,
            'physically_forced_sex_ever_id': 152,
            'physically_forced_sex_last_3months_id': 153,
            'seek_help_after_gbv_id': 158,
            'threatened_for_sexual_acts_ever_id': 156,
            'threatened_for_sexual_acts_last_3months_id': 157,
            'threats_to_hurt_ever_id': 144,
            'threats_to_hurt_last_3months_id': 145,
            'providers_sought': 159,
            'preferred_providers': 171,
        }
        for k, v in cols.items():
            if k == 'providers_sought':
                val = row.get(k)
                if val is not None:
                    sp = val.split(",")
                    for p in sp:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_gbv_sought_provider().get(p), value='Yes')
            elif k == 'preferred_providers':
                val = row.get(k)
                if val is not None:
                    sp = val.split(",")
                    for p in sp:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_gbv_preferred_provider().get(p), value='Yes')
            elif k == 'economic_threat_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'economic_threat_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'humiliated_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'humiliated_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'insulted_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'insulted_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'knowledge_of_gbv_help_centres_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physical_violence_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physical_violence_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_other_sex_acts_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_other_sex_acts_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_sex_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_sex_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'seek_help_after_gbv_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threatened_for_sexual_acts_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threatened_for_sexual_acts_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threats_to_hurt_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threats_to_hurt_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_education_and_employment(self, ws, i, row):
        cols = {
            'current_school_name': 66,
            'current_class': 70,
            'current_school_level_other': 69,
            'current_edu_supporter_list': 71,
            'current_education_supporter_other': 76,
            'reason_not_in_school_other': 78,
            'dropout_class': 80,
            'life_wish_other': 83,
            'current_income_source_other': 85,
            'banking_place_other': 88,
            'banking_place_id': 87,
            'current_income_source_id': 84,
            'current_school_level_id': 68,
            'current_school_type_id': 67,
            'currently_in_school_id': 65,
            'dropout_school_level_id': 81,
            'has_savings_id': 86,
            'last_time_in_school_id': 79,
            'life_wish_id': 82,
            'reason_not_in_school_id': 77
        }
        for k, v in cols.items():
            if k == 'current_edu_supporter_list':
                val = row.get(k)
                if val is not None:
                    sps = val.split(",")
                    for s in sps:
                        if s == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_education_supporter().get(s), value='Yes')
            elif k == 'banking_place_id':
                val = row.get(k)
                if val is not None:
                    item = self.bankingPlaceDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_income_source_id':
                val = row.get(k)
                if val is not None:
                    item = self.incomeSourceDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_school_level_id':
                val = row.get(k)
                if val is not None:
                    item = self.schoolLevelDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_school_type_id':
                val = row.get(k)
                if val is not None:
                    item = self.schoolTypeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_in_school_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'dropout_school_level_id':
                val = row.get(k)
                if val is not None:
                    item = self.schoolLevelDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_savings_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'last_time_in_school_id':
                val = row.get(k)
                if val is not None:
                    item = self.lastInSchoolDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'life_wish_id':
                val = row.get(k)
                if val is not None:
                    item = self.lifeWishDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'reason_not_in_school_id':
                val = row.get(k)
                if val is not None:
                    item = self.reasonNotInSchoolDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_hiv_testing(self, ws, i, row):
        cols = {
            'care_facility_enrolled': 93,
            'reason_not_in_hiv_care_other': 95,
            'reason_not_tested_for_hiv': 96,
            'reason_never_tested_for_hiv_other': 103,
            'enrolled_in_hiv_care_id': 92,
            'ever_tested_for_hiv_id': 89,
            'knowledge_of_hiv_test_centres_id': 106,
            'last_test_result_id': 91,
            'period_last_tested_id': 90,
            'reason_not_in_hiv_care_id': 94
        }
        for k, v in cols.items():
            if k == 'reason_not_tested_for_hiv':
                val = row.get(k)
                if val is not None:
                    rns = val.split(",")
                    for r in rns:
                        if r == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_reason_never_tested_for_hiv().get(r), value='Yes')
            elif k == 'enrolled_in_hiv_care_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_tested_for_hiv_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'knowledge_of_hiv_test_centres_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'last_test_result_id':
                val = row.get(k)
                if val is not None:
                    item = self.hivTestDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'period_last_tested_id':
                val = row.get(k)
                if val is not None:
                    item = self.periodDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'reason_not_in_hiv_care_id':
                val = row.get(k)
                if val is not None:
                    item = self.reasonNotInCareDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_dreams_programmes(self):
        return {
            '1': 195,
            '2': 196,
            '3': 197,
            '4': 198,
            '5': 199,
            '6': 200,
            '7': 201,
            '8': 202,
            '96': 203
        }

    def map_drugs(self):
        return {
            '1': 185,
            '2': 186,
            '3': 187,
            '4': 188,
            '5': 189,
            '6': 190,
            '7': 191,
            '8': 192,
            '96': 193

        }

    def map_gbv_preferred_provider(self):
        return {
            '1': 171,
            '2': 172,
            '3': 173,
            '4': 174,
            '5': 175,
            '6': 176,
            '7': 177,
            '8': 178,
            '9': 179,
            '10': 180,
            '96': 181

        }

    def map_gbv_sought_provider(self):
        return {
            '1': 159,
            '2': 160,
            '3': 161,
            '4': 162,
            '5': 163,
            '6': 164,
            '7': 165,
            '8': 166,
            '9': 167,
            '10': 168,
            '96': 169

        }

    def map_fp_method(self):
        return {
            '1': 130,
            '2': 131,
            '3': 132,
            '4': 133,
            '5': 134,
            '6': 135,
            '96': 136

        }

    def map_reason_never_tested_for_hiv(self):
        return {
            '1': 96,
            '2': 97,
            '3': 98,
            '4': 99,
            '5': 100,
            '6': 101,
            '7': 102,
            '8': 103,
            '9': 104,
            '96': 105

        }

    def map_education_supporter(self):
        return {
            '1': 71,
            '2': 72,
            '3': 73,
            '4': 74,
            '5': 75,
            '96': 76
        }

    def map_disability_type(self):
        return {
            '1': 51,
            '2': 52,
            '3': 53,
            '4': 54,
            '5': 55,
            '96': 56
        }

    def map_implementing_partner(self):
        return {
            1: 'Afya Jijini',
            2: 'AIHA',
            3: 'Aphiaplus Western',
            4: 'Global Communities',
            5: 'Henry Jackson Foundation',
            6: 'HWWK',
            7: 'IMC',
            8: 'IRDO',
            9: 'LVCT Health',
            10: 'Peace Corps'
        }

    def map_verification_document(self):
        return {
            4: 'Baptismal card',
            1: 'Birth Certificate',
            2: 'National ID',
            3: 'National ID waiting card',
        }

    def map_marital_status_codes(self):
        return {
            1: 'Single',
            2: 'Married / Cohabiting',
            3: 'Separated / Divorced',
            4: 'Widowed',
        }




    def headOfHouseHoldDictionary(self):
        return {
            1: "Self",
            2: "Father",
            3: "Mother",
            4: "Sibling",
            5: "Uncle/Aunt",
            6: "Grandparents",
            7: "Husband/Partner",
            96: "Other/Specify"
        }

    def yesNoDictionary(self):
        return {
            1: "YES",
            2: "NO",
            3: "Don't Know",
            4: "Not Applicable"
        }

    def floorMaterialDictionary(self):
        return {
            1: "Earth/Mud/Dung/Sand",
            2: "Wood planks",
            3: "Ceramic tiles",
            4: "Cement",
            96: "Other (Specify)"
        }

    def wallMaterialDictionary(self):
        return {
            1: "No Walls",
            2: "Dung/Mud",
            3: "Stone with mud",
            4: "Plywood/Cardboard",
            5: "Carton",
            6: "Wood",
            7: "Stone/Cement",
            96: "Other (Specify)"
        }

    def roofMaterialDictionary(self):
        return {
            1: "Grass/Thatch/Makuti",
            2: "Tin cans",
            3: "Corrugated iron sheet",
            4: "Asbestos sheets",
            5: "Concrete",
            96: "Other (Specify)"
        }

    def drinkingWaterSourceDictionary(self):
        return {
            1: "Piped water",
            2: "Open well",
            3: "Covered well/borehole",
            4: "Surface water (River, Spring and Lakes)",
            5: "Rain water",
            96: "Other (Specify)"
        }

    def frequencyDictionary(self):
        return {
                    1: "Rarely (1-2 days)",
            2: "Sometimes (3-10 days)",
            3: "Often (more than 10 days)",
            4: "Always",
            5: "Sometimes",
            6: "Never",
            7: "Often",
            8: "Not in the last 3 months",
            9: "Everyday",
            10: "5 to 6 times a week",
            11: "3 to 4 times a week",
            12: "Once a week",
            13: "2 to 3 times a month",
            14: "Once a month",
            15: "3 to 11 times in the past year",
            16: "1 to 2 times in the past year",
            96: "Other (Specify)"

        }

    def schoolTypeDictionary(self):
        return {
            1: "Formal",
            2: "Informal"
        }

    def schoolLevelDictionary(self):
        return {
            1: "Nursery",
            2: "PRIMARY_LEVEL",
            3: "SECONDAY_LEVEL",
            4: "TERTIARY_LEVEL",
            5: "VOCATIONAL_LEVEL",
            96: "Other (Specify)"
        }


    def reasonNotInSchoolDictionary(self):
        return {
            1: "Completed High School",
            2: "Lack of fees",
            3: "Pregnancy",
            4: "Peer pressure",
            5: "Not interested",
            6: "Awaiting to join secondary",
            7: "Got married",
            96: "Other (Specify)"
        }

    def lastInSchoolDictionary(self):
        return {
            1: "Less than 6 months ago",
            2: "6 to 12 months",
            3: "1to 2 years ago",
            4: "More than 2 years",
            5: "Never attended school",
            96: "Other (Specify)"
        }

    def lifeWishDictionary(self):
        return {
            1: "Pursue a course",
            2: "Start a business",
            3: "Go back to school",
            4: "Get married",
            96: "Other (Specify)"
        }

    def incomeSourceDictionary(self):
        return {
            1: "Formally employed",
            2: "Business person",
            3: "Casual labour",
            4: "Petty trade",
            5: "Farmer",
            6: "None",
            96: "Other (Specify)"
        }

    def bankingPlaceDictionary(self):
        return {
            1: "At home",
            2: "Table banking",
            3: "In the bank",
            96: "Other (Specify)"
        }

    def periodDictionary(self):
        return {
            1: "Less than 6 Months ago",
            2: "6 - 12 months ago",
            3: "1- 2 years ago",
            4: "More than 2 years ago",
            5: "Never attended to school",
            6: "Less than 3 months ago",
            7: "3-5 months ago",
            8: "6-12 months ago",
            9: "More than 12 months ago"
        }

    def hivTestDictionary(self):
        return {
            1: "Positive",
            2: "Negative",
            3: "Don't Know",
            4: "Declined to disclose"
        }

    def reasonNotInCareDictionary(self):
        return {
            1: "Facility is too far away",
            2: "I don't know where clinic is",
            3: "I can't afford it",
            4: "I feel healthy/not sick",
            5: "I fear people will know I have HIV if I go to clinic",
            6: "I feel I will be discriminated against by people at a facility",
            7: "The providers at facility are unfriendly",
            8: "I am taking alternative medicine that is not availble at the clinic",
            9: "I,m too busy to go",
            96: "Other(specify)"
        }

    def relativeAgeDictionary(self):
        return {
            2: "Same Age",
            3: "Younger",
            4: "Older"
        }

    def familyPlanningDictionary(self):
        return {
            1: "Pills",
            2: "Injectable",
            3: "Implants",
            4: "IUCD",
            5: "Condom",
            6: "Permanent (Tube Ligation)",
            96: "Other (Specify)"
        }

    def reasonNoInFPDictionary(self):
        return {
           1: "Not sexually active",
           2: "Religious reasons",
           3: "Can not afford",
           4: "Do not know where to get",
           5: "Currently pregnant",
           96: "Other (Specify)"
        }